from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import case, desc, func, select
from sqlalchemy.orm import Session

from models import MetricLog, PassengerObservation, SystemAlert


@dataclass(frozen=True)
class ShiftReportCsv:
    filename: str
    content: str


@dataclass(frozen=True)
class ShiftReportXlsx:
    filename: str
    content: bytes


LOCAL_TZ = ZoneInfo("Asia/Singapore")
REPORT_WINDOW_HOURS = 24
TIMELINE_SAMPLE_MINUTES = 5
REPORT_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

TITLE_FILL = PatternFill("solid", fgColor="082F49")
SUBTITLE_FILL = PatternFill("solid", fgColor="F8FAFC")
SECTION_FILL = PatternFill("solid", fgColor="0E7490")
HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
SAFE_FILL = PatternFill("solid", fgColor="DCFCE7")
WARNING_FILL = PatternFill("solid", fgColor="FEF3C7")
CRITICAL_FILL = PatternFill("solid", fgColor="FEE2E2")


def generate_shift_report_csv(
    db: Session,
    *,
    capacities: dict[str, int],
    run_id: str | None = None,
) -> ShiftReportCsv:
    generated_at = datetime.now(timezone.utc)
    window_start = _report_window_start(generated_at)
    selected_run_id = _resolve_run_id(db, run_id)
    metrics = _get_metrics_for_run(db, selected_run_id, since=window_start)
    alerts = _get_alerts_for_run(db, selected_run_id, since=window_start)
    demographics = _get_demographics_for_run(db, selected_run_id, since=window_start)
    sampled_metrics = _sample_metrics(metrics, minutes=TIMELINE_SAMPLE_MINUTES)

    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")

    writer.writerow(["CAG Passenger Monitoring Shift Report"])
    writer.writerow([])
    _write_summary(writer, generated_at, window_start, selected_run_id, metrics, alerts)
    writer.writerow([])
    _write_zone_summary(writer, metrics, capacities)
    writer.writerow([])
    _write_alert_log(writer, alerts)
    writer.writerow([])
    _write_demographics(writer, demographics)
    writer.writerow([])
    _write_metric_timeline(writer, sampled_metrics)

    return ShiftReportCsv(
        filename=_filename(selected_run_id, generated_at),
        content=buffer.getvalue(),
    )


def generate_shift_report_xlsx(
    db: Session,
    *,
    capacities: dict[str, int],
    run_id: str | None = None,
) -> ShiftReportXlsx:
    generated_at = datetime.now(timezone.utc)
    window_start = _report_window_start(generated_at)
    selected_run_id = _resolve_run_id(db, run_id)
    metrics = _get_metrics_for_run(db, selected_run_id, since=window_start)
    alerts = _get_alerts_for_run(db, selected_run_id, since=window_start)
    demographics = _get_demographics_for_run(db, selected_run_id, since=window_start)
    sampled_metrics = _sample_metrics(metrics, minutes=TIMELINE_SAMPLE_MINUTES)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Shift Report"
    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = False

    _style_columns(worksheet)
    row = 1
    row = _xlsx_title(worksheet, row, generated_at, window_start, selected_run_id)
    row = _xlsx_summary(worksheet, row, generated_at, window_start, selected_run_id, metrics, alerts)
    row = _xlsx_zone_summary(worksheet, row, metrics, capacities)
    row = _xlsx_alert_log(worksheet, row, alerts)
    row = _xlsx_demographics(worksheet, row, demographics)
    _xlsx_metric_timeline(worksheet, row, sampled_metrics)

    output = io.BytesIO()
    workbook.save(output)

    return ShiftReportXlsx(
        filename=_filename(selected_run_id, generated_at, extension="xlsx"),
        content=output.getvalue(),
    )


def _resolve_run_id(db: Session, requested_run_id: str | None) -> str | None:
    if requested_run_id:
        return requested_run_id

    statement = select(MetricLog.run_id).order_by(desc(MetricLog.timestamp), desc(MetricLog.id)).limit(1)
    return db.scalar(statement)


def _get_metrics_for_run(db: Session, run_id: str | None, *, since: datetime) -> list[MetricLog]:
    if not run_id:
        return []

    statement = (
        select(MetricLog)
        .where(MetricLog.run_id == run_id)
        .where(MetricLog.timestamp >= since)
        .order_by(MetricLog.timestamp, MetricLog.id)
    )
    return list(db.scalars(statement).all())


def _get_alerts_for_run(db: Session, run_id: str | None, *, since: datetime) -> list[SystemAlert]:
    if not run_id:
        return []

    statement = (
        select(SystemAlert)
        .where(SystemAlert.run_id == run_id)
        .where(SystemAlert.timestamp >= since)
        .order_by(SystemAlert.timestamp, SystemAlert.id)
    )
    return list(db.scalars(statement).all())


def _get_demographics_for_run(db: Session, run_id: str | None, *, since: datetime) -> dict[str, int]:
    if not run_id:
        return _empty_demographics()

    statement = (
        select(
            func.count(PassengerObservation.id).label("total_analyzed"),
            func.coalesce(func.sum(case((PassengerObservation.gender == "male", 1), else_=0)), 0).label("males"),
            func.coalesce(func.sum(case((PassengerObservation.gender == "female", 1), else_=0)), 0).label("females"),
            func.coalesce(
                func.sum(case((PassengerObservation.gender.not_in(["male", "female"]), 1), else_=0)),
                0,
            ).label("unknown"),
            func.coalesce(func.sum(case((PassengerObservation.age < 18, 1), else_=0)), 0).label("minors"),
        )
        .where(PassengerObservation.run_id == run_id)
        .where(PassengerObservation.timestamp >= since)
    )
    row = db.execute(statement).one()
    return {
        "total_analyzed": int(row.total_analyzed or 0),
        "males": int(row.males or 0),
        "females": int(row.females or 0),
        "unknown": int(row.unknown or 0),
        "minors": int(row.minors or 0),
    }


def _write_summary(
    writer: csv.writer,
    generated_at: datetime,
    window_start: datetime,
    run_id: str | None,
    metrics: list[MetricLog],
    alerts: list[SystemAlert],
) -> None:
    writer.writerow(["SECTION", "Summary"])
    writer.writerow(["Field", "Value"])
    writer.writerow(["Generated At", _format_datetime(generated_at)])
    writer.writerow(["Report Window Start", _format_datetime(window_start)])
    writer.writerow(["Report Window End", _format_datetime(generated_at)])
    writer.writerow(["Report Window", f"Latest {REPORT_WINDOW_HOURS} hours, Singapore time"])
    writer.writerow(["Run ID", run_id or "No metric data available"])

    if not metrics:
        writer.writerow(["Status", f"No metric data available in the latest {REPORT_WINDOW_HOURS} hours"])
        writer.writerow(["Total Metric Records", 0])
        writer.writerow(["Total Alerts", len(alerts)])
        return

    summary = _metric_summary(metrics)
    writer.writerow(["Start Time", _format_datetime(summary["start_time"])])
    writer.writerow(["End Time", _format_datetime(summary["end_time"])])
    writer.writerow(["Duration Minutes", summary["duration_minutes"]])
    writer.writerow(["Total Metric Records", len(metrics)])
    writer.writerow(["Latest Passenger Count", summary["latest_metric"].passenger_count])
    writer.writerow(["Peak Passenger Count", summary["peak_metric"].passenger_count])
    writer.writerow(["Peak Passenger Count Time", _format_datetime(summary["peak_metric"].timestamp)])
    writer.writerow(["Total Alerts", len(alerts)])


def _write_zone_summary(writer: csv.writer, metrics: list[MetricLog], capacities: dict[str, int]) -> None:
    writer.writerow(["SECTION", "Zone Capacity Summary"])
    writer.writerow(["Zone", "Capacity", "Peak Count", "Peak Used %", "Worst Status", "Latest Count", "Latest Used %"])

    rows = _zone_summary_rows(metrics, capacities)
    if not rows:
        writer.writerow(["No zone data available"])
        return

    for row in rows:
        writer.writerow(
            [
                row["zone_id"],
                row["capacity"] if row["capacity"] is not None else "",
                row["peak_count"],
                _format_percent(row["peak_percent"]),
                row["worst_status"].title(),
                row["latest_count"],
                _format_percent(row["latest_percent"]),
            ]
        )


def _write_alert_log(writer: csv.writer, alerts: list[SystemAlert]) -> None:
    writer.writerow(["SECTION", "Alert Log"])
    writer.writerow(["Timestamp", "Severity", "Message"])

    if not alerts:
        writer.writerow(["No alerts recorded"])
        return

    for alert in alerts:
        writer.writerow([_format_datetime(alert.timestamp), alert.severity, alert.message])


def _write_demographics(writer: csv.writer, demographics: dict[str, int]) -> None:
    writer.writerow(["SECTION", "Passenger Assistance Summary"])
    writer.writerow(["Total Analyzed", "Males", "Females", "Unknown", "Minors <18"])
    writer.writerow(
        [
            demographics["total_analyzed"],
            demographics["males"],
            demographics["females"],
            demographics["unknown"],
            demographics["minors"],
        ]
    )


def _write_metric_timeline(writer: csv.writer, metrics: list[MetricLog]) -> None:
    writer.writerow(["SECTION", f"Metric Timeline ({TIMELINE_SAMPLE_MINUTES}-minute samples)"])
    writer.writerow(["Timestamp", "Passenger Count", "Camera Online Count", "Zone Counts"])

    if not metrics:
        writer.writerow(["No metric data available"])
        return

    for metric in metrics:
        writer.writerow(_metric_timeline_values(metric))


def _xlsx_title(
    worksheet,
    row: int,
    generated_at: datetime,
    window_start: datetime,
    run_id: str | None,
) -> int:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = worksheet.cell(row=row, column=1, value="CAG Passenger Monitoring Shift Report")
    cell.fill = TITLE_FILL
    cell.font = Font(color="E0F2FE", bold=True, size=18)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[row].height = 30
    row += 1

    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = worksheet.cell(
        row=row,
        column=1,
        value=(
            f"Run ID: {run_id or 'No metric data available'} | "
            f"Generated: {_format_datetime(generated_at)} | "
            f"Window: {_format_datetime(window_start)} to {_format_datetime(generated_at)}"
        ),
    )
    cell.fill = SUBTITLE_FILL
    cell.font = Font(color="475569", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    return row


def _xlsx_summary(
    worksheet,
    row: int,
    generated_at: datetime,
    window_start: datetime,
    run_id: str | None,
    metrics: list[MetricLog],
    alerts: list[SystemAlert],
) -> int:
    row = _xlsx_section(worksheet, row, "Summary")
    row = _xlsx_header(worksheet, row, ["Field", "Value"])
    rows: list[list[Any]] = [
        ["Generated At", _format_datetime(generated_at)],
        ["Report Window Start", _format_datetime(window_start)],
        ["Report Window End", _format_datetime(generated_at)],
        ["Report Window", f"Latest {REPORT_WINDOW_HOURS} hours, Singapore time"],
        ["Run ID", run_id or "No metric data available"],
    ]

    if not metrics:
        rows.extend(
            [
                ["Status", f"No metric data available in the latest {REPORT_WINDOW_HOURS} hours"],
                ["Total Metric Records", 0],
                ["Total Alerts", len(alerts)],
            ]
        )
    else:
        summary = _metric_summary(metrics)
        rows.extend(
            [
                ["Start Time", _format_datetime(summary["start_time"])],
                ["End Time", _format_datetime(summary["end_time"])],
                ["Duration Minutes", summary["duration_minutes"]],
                ["Total Metric Records", len(metrics)],
                ["Latest Passenger Count", summary["latest_metric"].passenger_count],
                ["Peak Passenger Count", summary["peak_metric"].passenger_count],
                ["Peak Passenger Count Time", _format_datetime(summary["peak_metric"].timestamp)],
                ["Total Alerts", len(alerts)],
            ]
        )

    for values in rows:
        row = _xlsx_row(worksheet, row, values)
    return row + 1


def _xlsx_zone_summary(worksheet, row: int, metrics: list[MetricLog], capacities: dict[str, int]) -> int:
    row = _xlsx_section(worksheet, row, "Zone Capacity Summary")
    row = _xlsx_header(
        worksheet,
        row,
        ["Zone", "Capacity", "Peak Count", "Peak Used %", "Worst Status", "Latest Count", "Latest Used %"],
    )

    rows = _zone_summary_rows(metrics, capacities)
    if not rows:
        return _xlsx_row(worksheet, row, ["No zone data available"]) + 1

    for item in rows:
        row = _xlsx_row(
            worksheet,
            row,
            [
                item["zone_id"],
                item["capacity"] if item["capacity"] is not None else "",
                item["peak_count"],
                _format_percent(item["peak_percent"]),
                item["worst_status"].title(),
                item["latest_count"],
                _format_percent(item["latest_percent"]),
            ],
            status=item["worst_status"],
        )
    return row + 1


def _xlsx_alert_log(worksheet, row: int, alerts: list[SystemAlert]) -> int:
    row = _xlsx_section(worksheet, row, "Alert Log")
    row = _xlsx_header(worksheet, row, ["Timestamp", "Severity", "Message"])

    if not alerts:
        return _xlsx_row(worksheet, row, ["No alerts recorded"]) + 1

    for alert in alerts:
        row = _xlsx_row(
            worksheet,
            row,
            [_format_datetime(alert.timestamp), alert.severity, alert.message],
            status=(alert.severity or "").lower(),
        )
    return row + 1


def _xlsx_demographics(worksheet, row: int, demographics: dict[str, int]) -> int:
    row = _xlsx_section(worksheet, row, "Passenger Assistance Summary")
    row = _xlsx_header(worksheet, row, ["Total Analyzed", "Males", "Females", "Unknown", "Minors <18"])
    return (
        _xlsx_row(
            worksheet,
            row,
            [
                demographics["total_analyzed"],
                demographics["males"],
                demographics["females"],
                demographics["unknown"],
                demographics["minors"],
            ],
        )
        + 1
    )


def _xlsx_metric_timeline(worksheet, row: int, metrics: list[MetricLog]) -> int:
    row = _xlsx_section(worksheet, row, f"Metric Timeline ({TIMELINE_SAMPLE_MINUTES}-minute samples)")
    row = _xlsx_header(worksheet, row, ["Timestamp", "Passenger Count", "Camera Online Count", "Zone Counts"])

    if not metrics:
        return _xlsx_row(worksheet, row, ["No metric data available"])

    for metric in metrics:
        row = _xlsx_row(worksheet, row, _metric_timeline_values(metric))
    return row


def _xlsx_section(worksheet, row: int, label: str) -> int:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = worksheet.cell(row=row, column=1, value=label)
    cell.fill = SECTION_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    _style_row(worksheet, row)
    return row + 1


def _xlsx_header(worksheet, row: int, values: list[Any]) -> int:
    row = _xlsx_row(worksheet, row, values)
    for column in range(1, 8):
        cell = worksheet.cell(row=row - 1, column=column)
        cell.fill = HEADER_FILL
        cell.font = Font(color="0F172A", bold=True)
    return row


def _xlsx_row(worksheet, row: int, values: list[Any], *, status: str = "") -> int:
    padded = values[:7] + [""] * max(0, 7 - len(values))
    for column, value in enumerate(padded[:7], start=1):
        cell = worksheet.cell(row=row, column=column, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.alignment = Alignment(horizontal="right", vertical="top")

    fill = _status_fill(status)
    if fill is not None:
        for column in range(1, 8):
            worksheet.cell(row=row, column=column).fill = fill
            worksheet.cell(row=row, column=column).font = Font(bold=True)
    return row + 1


def _style_columns(worksheet) -> None:
    widths = {
        "A": 24,
        "B": 20,
        "C": 20,
        "D": 22,
        "E": 20,
        "F": 20,
        "G": 42,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _style_row(worksheet, row: int) -> None:
    for column in range(1, 8):
        worksheet.cell(row=row, column=column).border = THIN_BORDER


def _status_fill(status: str) -> PatternFill | None:
    normalized = status.lower()
    if normalized == "critical":
        return CRITICAL_FILL
    if normalized == "warning":
        return WARNING_FILL
    if normalized == "safe":
        return SAFE_FILL
    return None


def _metric_summary(metrics: list[MetricLog]) -> dict[str, Any]:
    start_time = metrics[0].timestamp
    end_time = metrics[-1].timestamp
    duration_minutes = round(max(0.0, (_as_utc(end_time) - _as_utc(start_time)).total_seconds()) / 60, 1)
    peak_metric = max(metrics, key=lambda metric: (metric.passenger_count, metric.timestamp, metric.id))
    latest_metric = metrics[-1]
    return {
        "start_time": start_time,
        "end_time": end_time,
        "duration_minutes": duration_minutes,
        "peak_metric": peak_metric,
        "latest_metric": latest_metric,
    }


def _zone_summary_rows(metrics: list[MetricLog], capacities: dict[str, int]) -> list[dict[str, Any]]:
    if not metrics:
        return []

    zones = sorted(set(capacities) | _zone_ids_from_metrics(metrics))
    if not zones:
        return []

    latest_counts = _parse_zone_counts(metrics[-1].zone_counts)
    rows: list[dict[str, Any]] = []
    for zone_id in zones:
        counts = [_zone_count(metric, zone_id) for metric in metrics]
        peak_count = max(counts) if counts else 0
        latest_count = _coerce_zone_count(latest_counts.get(zone_id)) or 0
        capacity = capacities.get(zone_id)
        peak_percent, worst_status = _capacity_status(peak_count, capacity)
        latest_percent, _ = _capacity_status(latest_count, capacity)
        rows.append(
            {
                "zone_id": zone_id,
                "capacity": capacity,
                "peak_count": peak_count,
                "peak_percent": peak_percent,
                "worst_status": worst_status,
                "latest_count": latest_count,
                "latest_percent": latest_percent,
            }
        )
    return rows


def _sample_metrics(metrics: list[MetricLog], *, minutes: int) -> list[MetricLog]:
    buckets: dict[datetime, MetricLog] = {}
    for metric in metrics:
        local_timestamp = _as_local(metric.timestamp)
        bucket_minute = local_timestamp.minute - (local_timestamp.minute % minutes)
        bucket = local_timestamp.replace(minute=bucket_minute, second=0, microsecond=0)
        buckets[bucket] = metric
    return list(buckets.values())


def _metric_timeline_values(metric: MetricLog) -> list[Any]:
    return [
        _format_datetime(metric.timestamp),
        metric.passenger_count,
        metric.camera_online_count if metric.camera_online_count is not None else "",
        metric.zone_counts or "",
    ]


def _empty_demographics() -> dict[str, int]:
    return {
        "total_analyzed": 0,
        "males": 0,
        "females": 0,
        "unknown": 0,
        "minors": 0,
    }


def _parse_zone_counts(value: str | None) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _zone_ids_from_metrics(metrics: list[MetricLog]) -> set[str]:
    zone_ids: set[str] = set()
    for metric in metrics:
        zone_ids.update(str(zone_id) for zone_id in _parse_zone_counts(metric.zone_counts))
    return zone_ids


def _zone_count(metric: MetricLog, zone_id: str) -> int:
    return _coerce_zone_count(_parse_zone_counts(metric.zone_counts).get(zone_id)) or 0


def _coerce_zone_count(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return max(0, int(value))
    if isinstance(value, dict):
        nested_counts = [_coerce_zone_count(item) for item in value.values()]
        valid_counts = [count for count in nested_counts if count is not None]
        return sum(valid_counts) if valid_counts else None
    return None


def _capacity_status(count: int, capacity: int | None) -> tuple[float | None, str]:
    if not capacity or capacity <= 0:
        return None, "unknown"

    percent_used = round((count / capacity) * 100, 1)
    if percent_used >= 85:
        return percent_used, "critical"
    if percent_used >= 60:
        return percent_used, "warning"
    return percent_used, "safe"


def _format_percent(value: float | None) -> str:
    return "" if value is None else f"{value:.1f}"


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return _as_local(value).strftime("%Y-%m-%d %H:%M:%S SGT")


def _report_window_start(generated_at: datetime) -> datetime:
    return (_as_local(generated_at) - timedelta(hours=REPORT_WINDOW_HOURS)).astimezone(timezone.utc)


def _as_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _as_local(value: datetime) -> datetime:
    return _as_utc(value).astimezone(LOCAL_TZ)


def _filename(run_id: str | None, generated_at: datetime, extension: str = "csv") -> str:
    safe_run_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", run_id or "no_data").strip("._-") or "no_data"
    return f"cag_shift_report_{safe_run_id}_{_as_local(generated_at).date().isoformat()}.{extension}"
